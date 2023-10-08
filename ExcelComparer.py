import openpyxl
import itertools

# Final変数
LIMIT = 50
inputs = ["n", "o", "p", "d"]
files = ["new", "old", "past"]
regions = ["Americas", "EMEA", "Pacific"]


print("No duplicate input allowed\n")
t = ""

# new, old, pastのどれをどれと比較するか入力を得る
while True:
    f = input("Compare (n/o/p) or Default(d): ")
    if f in inputs:
        inputs.remove(f)
        break
    
while f != "d":
    t = input("With (n/o/p): ")
    if t in inputs:
        break

# ファイル名に直す
if f == "d":
    f = "old"
    t = "new"
else:
    for i in files:
        if i[0] == f:
            f = i
        if i[0] == t:
            t = i

t = t + ".xlsx"
f = f + ".xlsx"

############################## newとoldの相違点を検索 ##############################

# 比較するExcelファイルを読み込む
wb1 = openpyxl.load_workbook(t)
wb2 = openpyxl.load_workbook(f)

# 比較するシートを選択する
sheet1 = wb1.active
sheet2 = wb2.active

# 2つのシートのセルを比較する
n = 0
print(f"\nFirst {LIMIT} changes...")

f1 = list(sheet1.iter_rows())
f2 = list(sheet2.iter_rows())
f1break = []
f2break = []

# "Last Update:"の位置を確認
n = 0
for row in f1:
    if row[0].value == "Last Update:":
        f1break.append(n)
    n += 1

n = 0
for row in f2:
    if row[0].value == "Last Update:":
        f2break.append(n)
    n += 1

# リストスライスの都合上追加
f1break.append(10000)
f2break.append(10000)

n = 0
for i in range(len(regions)):
    print("\n" + regions[i] + "\n")
    for row1, row2 in zip(f1[f1break[i]:f1break[i + 1]], f2[f2break[i]:f2break[i + 1]]):
        for cell1, cell2 in zip(row1, row2):
            if n == LIMIT:
                break

            if cell1.value != cell2.value:
                print(f"{cell2.coordinate} {cell1.coordinate}: {cell2.value} => {cell1.value}")
                n += 1
    if not n:
        print("No Changes")
    n = 0



