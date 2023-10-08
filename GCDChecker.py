import requests
import openpyxl
from bs4 import BeautifulSoup

# 変更箇所の表示上限
LIMIT = 30

########################## newをoldに、oldをpastに上書き ##########################

# Excelファイルを読み込む
wb_new = openpyxl.load_workbook("new.xlsx")
wb_old = openpyxl.load_workbook("old.xlsx")
# 相違点がない場合pastを元に戻す
wb_past = openpyxl.load_workbook("past.xlsx")

# 上書き
wb_old.save("past.xlsx")
wb_new.save("old.xlsx")

########################### new.xlsxに最新の GCDをコピー ###########################

# WebページのURL
url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRmmWiBmMMD43m5VtZq54nKlmj0ZtythsA1qCpegwx-iRptx2HEsG0T3cQlG1r2AIiKxBWnaurJZQ9Q/pubhtml#"

# WebページのHTMLを取得する
response = requests.get(url)
html = response.text

# BeautifulSoupを使用してHTMLを解析する
soup = BeautifulSoup(html, "html.parser")

# td要素の中身のみ抽出する
elements = soup.find_all('td')
contents = [element.text.strip() for element in elements]

# Excelファイルを開く
wb = openpyxl.Workbook()
ws = wb.active

# リストの要素をExcelに書き込む
n = 11  # n要素ごとに新しい行に書き込む
for i in range(0, len(contents), n):
    row = [(contents[j] if j < len(contents) else '') for j in range(i, i+n)]
    ws.append(row)

# Excelファイルを保存する
wb.save("new.xlsx")
print("Saved new GCD...")

############################## newとoldの相違点を検索 ##############################

# 比較するExcelファイルを読み込む
wb_new = openpyxl.load_workbook("new.xlsx")
wb_old = openpyxl.load_workbook("old.xlsx")

# 比較するシートを選択する
sheet_new = wb_new.active
sheet_old = wb_old.active

# 2つのシートのセルを比較する
n = 0
print(f"\nFirst {LIMIT} changes...")
for row_new, row_old in zip(sheet_new.iter_rows(), sheet_old.iter_rows()):
    for cell_new, cell_old in zip(row_new, row_old):
        if n == LIMIT:
            break
        if cell_new.value != cell_old.value:
            print(f"{cell_old.coordinate}: {cell_old.value} => {cell_new.value}")
            n += 1
            
# 相違点がない場合pastを元に戻す
if not n:
    # Excelファイルを読み込む
    wb_old = openpyxl.load_workbook("old.xlsx")
    wb_past2 = openpyxl.load_workbook("past.xlsx")

    # 上書き
    wb_old.save("new.xlsx")
    wb_past2.save("old.xlsx")
    wb_past.save("past.xlsx")
    
    print("\nNo changes found...\n\nInitial state restored")






